"""仪器台账路由"""
from io import BytesIO
from datetime import datetime, date
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from app.database import get_db
from app.models import Instrument, User
from app.schemas import InstrumentCreate, InstrumentUpdate, ImportResult
from app.utils.auth import get_current_user, require_role, apply_department_filter
from app.utils.audit import log

router = APIRouter()

def _auto_code(db: Session) -> str:
    import uuid
    today = date.today().strftime("%Y%m%d")
    return f"INS-{today}-{uuid.uuid4().hex[:6].upper()}"

# ── 列表（含搜索、分页） ──
@router.get("/")
def list_instruments(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: str = Query(None, description="关键词搜索（fields JSONB 内全文匹配）"),
    department_id: int = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(Instrument)
    query = apply_department_filter(query, Instrument, current_user)
    if department_id:
        query = query.filter(Instrument.department_id == department_id)
    if search:
        query = query.filter(Instrument.fields.cast(String).ilike(f"%{search}%"))
    from sqlalchemy import String
    total = query.count()
    items = query.order_by(Instrument.updated_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {"data": _serialize_instruments(items, db), "meta": {"total": total, "page": page, "page_size": page_size}}

# ── 详情 ──
@router.get("/{instrument_id}")
def get_instrument(instrument_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    inst = db.query(Instrument).filter(Instrument.id == instrument_id).first()
    if not inst:
        raise HTTPException(404, detail="仪器不存在")
    return {"data": _serialize_one(inst, db)}

# ── 新增 ──
@router.post("/", status_code=201)
def create_instrument(body: InstrumentCreate, db: Session = Depends(get_db),
                      current_user: User = Depends(require_role(["admin", "system_manager", "dept_measurer"]))):
    inst = Instrument(code=_auto_code(db), department_id=body.department_id or current_user.department_id, fields=body.fields or {})
    db.add(inst)
    db.flush()
    log(db, current_user.id, "create", "instrument", inst.id, summary=f"新增仪器 {inst.code}")
    db.commit()
    db.refresh(inst)
    return {"data": _serialize_one(inst, db)}

# ── 更新 ──
@router.put("/{instrument_id}")
def update_instrument(instrument_id: int, body: InstrumentUpdate, db: Session = Depends(get_db),
                      current_user: User = Depends(require_role(["admin", "system_manager", "dept_measurer"]))):
    inst = db.query(Instrument).filter(Instrument.id == instrument_id).first()
    if not inst:
        raise HTTPException(404, detail="仪器不存在")
    if body.department_id is not None:
        inst.department_id = body.department_id
    if body.fields is not None:
        old_fields = dict(inst.fields or {})
        inst.fields = body.fields
        log(db, current_user.id, "update", "instrument", inst.id, summary=f"更新仪器 {inst.code}", changes={"fields": {"old": old_fields, "new": body.fields}})
    db.commit()
    db.refresh(inst)
    return {"data": _serialize_one(inst, db)}

# ── 删除 ──
@router.delete("/{instrument_id}")
def delete_instrument(instrument_id: int, db: Session = Depends(get_db),
                      current_user: User = Depends(require_role(["admin"]))):
    inst = db.query(Instrument).filter(Instrument.id == instrument_id).first()
    if not inst:
        raise HTTPException(404, detail="仪器不存在")
    log(db, current_user.id, "delete", "instrument", inst.id, summary=f"删除仪器 {inst.code}")
    db.delete(inst)
    db.commit()
    return {"data": None}

# ── 清空 ──
@router.delete("/batch/clear")
def clear_all(db: Session = Depends(get_db), current_user: User = Depends(require_role(["admin"]))):
    deleted = db.query(Instrument).delete()
    log(db, current_user.id, "delete", "instrument", summary=f"清空全部仪器，共 {deleted} 条")
    db.commit()
    return {"data": {"deleted_count": deleted}}

# ── 导入 ──
@router.post("/import")
def import_instruments(file: UploadFile = File(...), db: Session = Depends(get_db),
                       current_user: User = Depends(require_role(["admin", "system_manager"]))):
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(400, detail="仅支持 .xlsx / .xls")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(file.file.read()))
    except Exception as e:
        raise HTTPException(400, detail=f"无法解析: {str(e)}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, detail="文件为空")
    headers = [str(h).strip() if h else f"列{i+1}" for i, h in enumerate(rows[0])]
    success = 0
    batch = []
    for row in rows[1:]:
        fields = {}
        for i, h in enumerate(headers):
            v = row[i] if i < len(row) else None
            if v is None: fields[h] = ""
            elif isinstance(v, datetime): fields[h] = v.strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(v, date): fields[h] = v.strftime("%Y-%m-%d")
            else: fields[h] = str(v).strip()
        if all(not v for v in fields.values()):
            continue
        batch.append(Instrument(code=_auto_code(db), department_id=current_user.department_id, fields=fields))
        success += 1
        if len(batch) >= 100:
            db.add_all(batch); db.commit(); batch = []
    if batch:
        db.add_all(batch); db.commit()
    log(db, current_user.id, "import", "instrument", summary=f"导入 {success} 条仪器")
    return {"data": ImportResult(total_rows=len(rows)-1, success_count=success, failure_count=0, errors=[], message=f"导入完成：成功 {success} 条")}

# ── 导出 ──
@router.get("/export/excel")
def export_instruments(search: str = Query(None), db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user)):
    query = apply_department_filter(db.query(Instrument), Instrument, current_user)
    if search:
        from sqlalchemy import String
        query = query.filter(Instrument.fields.cast(String).ilike(f"%{search}%"))
    items = query.all()
    # Build all keys
    all_keys = []
    seen = set()
    for inst in items:
        for k in inst.fields or {}:
            if k not in seen:
                all_keys.append(k)
                seen.add(k)
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["仪器编号"] + all_keys)
    for inst in items:
        row = [inst.code] + [str(inst.fields.get(k, "")) for k in all_keys]
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    from fastapi.responses import StreamingResponse
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=instruments.xlsx"})

def _serialize_instruments(items, db):
    depts = {}
    return [_serialize_one(inst, db, depts) for inst in items]

def _serialize_one(inst, db, depts=None):
    dept_name = None
    if inst.department_id:
        if depts is not None:
            if inst.department_id not in depts:
                from app.models import Department
                d = db.query(Department).filter(Department.id == inst.department_id).first()
                depts[inst.department_id] = d.name if d else None
            dept_name = depts[inst.department_id]
        elif inst.department:
            dept_name = inst.department.name
    return {
        "id": inst.id, "code": inst.code, "department_id": inst.department_id,
        "department_name": dept_name, "fields": inst.fields or {},
        "created_at": str(inst.created_at) if inst.created_at else None,
        "updated_at": str(inst.updated_at) if inst.updated_at else None,
    }
