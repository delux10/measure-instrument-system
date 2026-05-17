"""Tests for Excel batch import functionality."""
import pytest
import io
from datetime import date
from unittest.mock import MagicMock

import openpyxl

from app.services.instrument_service import (
    _parse_date,
    validate_row,
    build_instrument_from_values,
    MES_COLUMN_MAP,
    STATUS_MAP,
)
from app.schemas.instrument import ImportRowError


def _make_excel_bytes(headers: list, rows: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── _parse_date ──────────────────────────────────────────────

@pytest.mark.parametrize("val,expected", [
    ("2024-01-15", date(2024, 1, 15)),
    ("2024/01/15", date(2024, 1, 15)),
    ("2024.01.15", date(2024, 1, 15)),
    ("2024年1月15日", date(2024, 1, 15)),
    ("01/15/2024", date(2024, 1, 15)),
])
def test_parse_date_valid_formats(val, expected):
    assert _parse_date(val) == expected


def test_parse_date_invalid_returns_none():
    assert _parse_date("not a date") is None
    assert _parse_date("") is None


# ── validate_row ────────────────────────────────────────────

def test_validate_row_missing_code():
    errors = validate_row({"name": "卡尺", "model": "A1", "category_name": "长度类"}, set(), {}, {}, {})
    assert len(errors) >= 1
    assert "仪器编号" in errors[0].field


def test_validate_row_missing_name():
    errors = validate_row({"code": "M001", "model": "A1", "category_name": "长度类"}, set(), {}, {}, {})
    assert any("名称" in e.field for e in errors)


def test_validate_row_missing_model():
    errors = validate_row({"code": "M001", "name": "卡尺", "category_name": "长度类"}, set(), {}, {}, {})
    assert any("型号" in e.field for e in errors)


def test_validate_row_missing_category():
    errors = validate_row({"code": "M001", "name": "卡尺", "model": "A1"}, set(), {}, {}, {})
    assert any("分类" in e.field for e in errors)


def test_validate_row_duplicate_code():
    errors = validate_row(
        {"code": "M001", "name": "卡尺", "model": "A1", "category_name": "长度类"},
        {"M001"},
        {"长度类": 1},
        {},
        {}
    )
    assert any("已存在" in e.message for e in errors)


def test_validate_row_category_not_found():
    errors = validate_row(
        {"code": "M001", "name": "卡尺", "model": "A1", "category_name": "不存在的分类"},
        set(),
        {},
        {},
        {}
    )
    assert any("不存在" in e.message for e in errors)


def test_validate_row_category_found_by_path():
    errors = validate_row(
        {"code": "M001", "name": "卡尺", "model": "A1", "category_name": "长度类/卡尺"},
        set(),
        {},
        {"长度类/卡尺": 2},
        {}
    )
    assert len(errors) == 0


def test_validate_row_all_required_present():
    errors = validate_row(
        {"code": "M001", "name": "卡尺", "model": "A1", "category_name": "长度类"},
        set(),
        {"长度类": 1},
        {},
        {}
    )
    assert len(errors) == 0


def test_validate_row_department_not_found_soft_fail():
    errors = validate_row(
        {"code": "M001", "name": "卡尺", "model": "A1", "category_name": "长度类", "department_name": "不存在部门"},
        set(),
        {"长度类": 1},
        {},
        {}
    )
    assert any("部门" in (e.field or "") for e in errors)


def test_validate_row_invalid_date():
    errors = validate_row(
        {"code": "M001", "name": "卡尺", "model": "A1", "category_name": "长度类", "manufacture_date": "abc"},
        set(),
        {"长度类": 1},
        {},
        {}
    )
    assert any(e.field == "manufacture_date" for e in errors)


def test_validate_row_invalid_price():
    errors = validate_row(
        {"code": "M001", "name": "卡尺", "model": "A1", "category_name": "长度类", "price": "abc"},
        set(),
        {"长度类": 1},
        {},
        {}
    )
    assert any(e.field == "price" for e in errors)


# ── build_instrument_from_values ────────────────────────────

def test_build_basic_instrument():
    values = {"code": "M001", "name": "卡尺", "model": "A1"}
    inst = build_instrument_from_values(values.copy(), {}, {"长度类": 1}, {})
    assert inst.code == "M001"
    assert inst.name == "卡尺"
    assert inst.model == "A1"
    assert inst.status == "in_use"


def test_build_with_category():
    values = {"code": "M002", "name": "天平", "model": "B2", "category_name": "力学类"}
    inst = build_instrument_from_values(values.copy(), {}, {"力学类": 2}, {})
    assert inst.category_id == 2


def test_build_with_dept_name_to_id():
    values = {"code": "M003", "name": "X", "model": "Y", "department_name": "质管科"}
    inst = build_instrument_from_values(values.copy(), {"质管科": 5}, {}, {})
    assert inst.department_id == 5


def test_build_maps_status():
    values = {"code": "M004", "name": "X", "model": "Y", "status": "停用"}
    inst = build_instrument_from_values(values.copy(), {}, {}, {})
    assert inst.status == "stopped"


def test_build_parses_calibration_cycle():
    values = {"code": "M005", "name": "X", "model": "Y", "calibration_cycle": "12个月"}
    inst = build_instrument_from_values(values.copy(), {}, {}, {})
    assert inst.calibration_cycle == 12


def test_build_parses_price():
    values = {"code": "M006", "name": "X", "model": "Y", "price": "1,234.56"}
    inst = build_instrument_from_values(values.copy(), {}, {}, {})
    assert inst.price == 1234.56


def test_build_computes_next_cal_date():
    values = {
        "code": "M007", "name": "X", "model": "Y",
        "last_cal_date": "2024-01-15",
        "calibration_cycle": 12,
    }
    inst = build_instrument_from_values(values.copy(), {}, {}, {})
    assert inst.next_cal_date == date(2025, 1, 15)


# ── MES_COLUMN_MAP completeness ────────────────────────────

def test_column_map_contains_required_fields():
    """Verify the column map covers the 4 required import fields."""
    values = set(MES_COLUMN_MAP.values())
    assert "code" in values
    assert "name" in values
    assert "model" in values
    assert "category_name" in values


# ── Excel generation helper ─────────────────────────────────

def test_make_excel_bytes_produces_valid_xlsx():
    content = _make_excel_bytes(
        ["仪器编号", "仪器名称", "型号规格", "仪器分类"],
        [["M001", "卡尺", "A1", "长度类"]]
    )
    wb = openpyxl.load_workbook(io.BytesIO(content))
    assert len(list(wb.active.iter_rows(values_only=True))) == 2


# ── FastAPI integration test ────────────────────────────────

from fastapi.testclient import TestClient
from app.main import app
from app.database import get_db
from unittest.mock import patch

client = TestClient(app)


def test_import_endpoint_requires_auth():
    resp = client.post("/api/instruments/import")
    assert resp.status_code in (401, 403)


def test_import_health_check():
    resp = client.get("/api/health")
    assert resp.status_code == 200
    assert resp.json()["status"] == "ok"
